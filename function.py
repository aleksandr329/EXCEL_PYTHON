import openpyxl
from openpyxl import load_workbook
from constants import time_now, vopros1



def scan():  #<- Функция для сканирования и вывода информации в консоль и в файл
    file = 'Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'I{i}'].value is None:
            print('Программа завершила работу!')
            break

        if sheet[f'I{i}'].value in 'Не оплачен':
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            info = f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo} \n'''
            print(info)
            with open(f'Отчет {time_now}.txt', 'a') as file_txt:
                file_txt.write(info)

def scan2():  #<- Функция для определения пустой клетки для новой записи
    file = 'Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'A{i}'].value is None:
            return i
            break

def zapis():  #<-  Функция для записи данных
    file = 'Таблица учета Оплат.xlsx'
    wb = load_workbook(file)
    ws = wb['Лист1']
    number = scan2()
    ws['A' + str(number)] = number - 1
    data_scheta = input('Введите дату счета: ')
    ws['B' + str(number)] = data_scheta
    nomer_scheta = input('Введите номер счета: ')
    ws['C' + str(number)] = nomer_scheta
    name = input('Введите имя организации: ')
    ws['D' + str(number)] = name
    inn = int(input('Введите инн организации: '))
    ws['E' + str(number)] = inn
    vsa_summa = input('Введите всю сумму платежа: ')
    ws['F' + str(number)] = vsa_summa
    ostatoc = str(input('Введите остаток сколько нужно доплатить: '))
    ws['G' + str(number)] = ostatoc
    data_oplati = input('Введите дату до которой нужно внести остаток суммы: ')
    ws['H' + str(number)] = data_oplati
    status = input('Введите статус Оплачен или Не оплачен: ')
    ws['I' + str(number)] = status
    wb.save(file)
    wb.close()

def scan3(name):  #<- Функция для скана по имени фирмы
    file = 'Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'D{i}'].value is None:
            print('Программа завершила работу!')
            break

        if sheet[f'D{i}'].value.lower() in name.lower():
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            info = f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo} \n'''
            print(info)
            with open(f'Отчет по названию фирмы {time_now}.txt', 'a') as file_txt:
                file_txt.write(info)
            continue

        else:
            name_tabl = sheet[f'D{i}'].value.lower()
            name_user = name.lower()
            try:
                if name_tabl[0] in name_user[0] and name_tabl[1] in name_user[1] and name_tabl[2] in name_user[2]:
                    i_d = sheet[f'A{i}'].value
                    number = sheet[f'C{i}'].value
                    organizace = sheet[f'D{i}'].value
                    inn = sheet[f'E{i}'].value
                    summa = sheet[f'F{i}'].value
                    ostatok = sheet[f'G{i}'].value
                    chislo = sheet[f'H{i}'].value
                    info = f'''ID {i_d}, Номер счета: {number},
                    ИНН: {inn}, Организация: {organizace},
                    Общая сумма: {summa}, Остаток: {ostatok},
                    Оплатить до: {chislo} \n'''
                    print(info)
                    with open(f'Отчет по названию фирмы {time_now}.txt', 'a') as file_txt:
                        file_txt.write(info)
                    continue
            
                if name_tabl[0] in name_user[0] and name_tabl[1] in name_user[1] and name_tabl[2] in name_user[2] and name_tabl[3] in name_user[3]:
                    i_d = sheet[f'A{i}'].value
                    number = sheet[f'C{i}'].value
                    organizace = sheet[f'D{i}'].value
                    inn = sheet[f'E{i}'].value
                    summa = sheet[f'F{i}'].value
                    ostatok = sheet[f'G{i}'].value
                    chislo = sheet[f'H{i}'].value
                    info = f'''ID {i_d}, Номер счета: {number},
                    ИНН: {inn}, Организация: {organizace},
                    Общая сумма: {summa}, Остаток: {ostatok},
                    Оплатить до: {chislo} \n'''
                    print(info)
                    with open(f'Отчет по названию фирмы {time_now}.txt', 'a') as file_txt:
                        file_txt.write(info)
                    continue

            except IndexError:
                print('Вы ввели недостаточно символов чтобы определить название фирмы')
                break

        
def scan4(inn):  #<- Функция для скана по инн фирмы
    file = 'Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'E{i}'].value is None:
            print('Программа завершила работу!')
            break

        if sheet[f'E{i}'].value == inn:
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            status = sheet[f'I{i}'].value
            info = f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo}, Статус: {status} \n'''
            print(info)
            with open(f'Отчет по ИНН фирмы {time_now}.txt', 'a') as file_txt:
                file_txt.write(info)

def scan5(number):  #<- Функция для скана по номеру счета
    file = 'Таблица учета Оплат.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=file)
    wb.active = 0
    sheet = wb.active
    for i in range(2, 10000):
        if sheet[f'C{i}'].value is None:
            print('Программа завершила работу!')
            break

        if sheet[f'C{i}'].value.lower() in number.lower():
            i_d = sheet[f'A{i}'].value
            number = sheet[f'C{i}'].value
            organizace = sheet[f'D{i}'].value
            inn = sheet[f'E{i}'].value
            summa = sheet[f'F{i}'].value
            ostatok = sheet[f'G{i}'].value
            chislo = sheet[f'H{i}'].value
            print(f'''ID {i_d}, Номер счета: {number},
            ИНН: {inn}, Организация: {organizace},
            Общая сумма: {summa}, Остаток: {ostatok},
            Оплатить до: {chislo}''')
            vopros = input(vopros1)
            if vopros.lower() in 'статус':
                print('Оплачен или Не оплачен')
                vopros2 = input('Какой поставим статус? ')

                if vopros2.capitalize() in 'Оплачен':
                    wb = load_workbook(file)
                    ws = wb['Лист1']
                    ws['I' + str(i)] = vopros2.capitalize()
                    wb.save(file)
                    wb.close()

                if vopros2.capitalize() in 'Не оплачен':
                    wb = load_workbook(file)
                    ws = wb['Лист1']
                    ws['I' + str(i)] = vopros2.capitalize()
                    wb.save(file)
                    wb.close()
                    
            if vopros.lower() in 'остаток':
                ostatok_new = input('Какой остаток нужно будет доплатить? ')
                wb = load_workbook(file)
                ws = wb['Лист1']
                ws['G' + str(i)] = ostatok_new
                wb.save(file)
                wb.close()

            if vopros.lower() in 'дату':
                data_new = input('Какую дату поставить? ')
                wb = load_workbook(file)
                ws = wb['Лист1']
                ws['H' + str(i)] = data_new
                wb.save(file)
                wb.close()
